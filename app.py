# -*- coding: utf-8 -*-
import os
import io
import uuid
from datetime import datetime, date

from flask import Flask, request, jsonify, render_template, Response
from flask_cors import CORS

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from models import db, Evaluation, ShiftSession, Fine, Rep, CortexRequest
import data as D
import cortex as C

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL", f"sqlite:///{os.path.join(BASE_DIR, 'bankai.db')}"
).replace("postgres://", "postgresql://", 1)  # Railway بيدي postgres:// أحيانًا
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)

# CORS مفعّل بس على مسارات /api/cortex/*، عشان الـ Tampermonkey script
# (شغال على logistics.amazon.eg) يقدر يرفع الملف مباشرة من هناك.
# باقي الـ API متعمدين نسيبه من غير CORS.
CORS(
    app,
    resources={r"/api/cortex/*": {"origins": "https://logistics.amazon.eg"}},
)


def get_supervisor(branch, supervisor_id):
    """يرجع بيانات المشرف (بما فيها shift_id الثابت بتاعه) أو None."""
    b = D.BRANCHES.get(branch)
    if not b:
        return None
    for s in b["supervisors"]:
        if s["id"] == supervisor_id:
            return s
    return None


with app.app_context():
    db.create_all()


# ============================================================
# Pages
# ============================================================
@app.route("/")
def index():
    return render_template("index.html")


# ============================================================
# Static config data (branches, supervisors, shifts)
# ============================================================
@app.route("/api/branches")
def api_branches():
    return jsonify({
        "branches": {
            bid: {"label": b["label"], "supervisors": b["supervisors"]}
            for bid, b in D.BRANCHES.items()
        },
        "supervisor_shifts": D.SUPERVISOR_SHIFTS,
    })


# ============================================================
# Reps management (إضافة مندوبين من الواجهة)
# ============================================================
@app.route("/api/manage_reps")
def api_manage_reps_list():
    branch = request.args.get("branch", "")
    q = Rep.query.filter_by(active=True)
    if branch:
        q = q.filter_by(branch=branch)
    reps = q.order_by(Rep.created_at.desc()).all()
    return jsonify({"reps": [r.to_dict() for r in reps]})


@app.route("/api/manage_reps", methods=["POST"])
def api_manage_reps_add():
    payload = request.get_json(force=True) or {}
    required = ["name", "branch", "rep_type", "start", "end"]
    missing = [k for k in required if k not in payload]
    if missing:
        return jsonify({"error": f"missing: {missing}"}), 400

    name = str(payload["name"]).strip()
    branch = payload["branch"]
    rep_type = payload["rep_type"]
    email = (payload.get("email") or "").strip() or None

    if not name:
        return jsonify({"error": "invalid_name"}), 400
    if branch not in D.BRANCHES:
        return jsonify({"error": "invalid_branch"}), 400
    if rep_type not in ("full", "part"):
        return jsonify({"error": "invalid_type"}), 400
    try:
        start = int(payload["start"])
        end = int(payload["end"])
    except (TypeError, ValueError):
        return jsonify({"error": "invalid_hours"}), 400
    if not (0 <= start <= 23) or not (0 <= end <= 23):
        return jsonify({"error": "invalid_hours"}), 400

    rep = Rep(name=name, email=email, branch=branch, rep_type=rep_type, start=start, end=end)
    db.session.add(rep)
    db.session.commit()
    return jsonify({"ok": True, "rep": rep.to_dict()})


@app.route("/api/manage_reps/<int:rep_id>", methods=["DELETE"])
def api_manage_reps_delete(rep_id):
    rep = Rep.query.get(rep_id)
    if rep:
        rep.active = False
        db.session.commit()
    return jsonify({"ok": True})


# ============================================================
# Reps overlapping a supervisor's shift, merged with saved evaluations
# ============================================================
@app.route("/api/reps")
def api_reps():
    branch = request.args.get("branch", "")
    supervisor_id = request.args.get("supervisor_id", "")
    supervisor_shift_id = request.args.get("shift_id", "")
    day = request.args.get("day", "")

    if not all([branch, supervisor_id, supervisor_shift_id, day]):
        return jsonify({"error": "missing params"}), 400

    reps = D.get_overlapping_reps(
        branch, supervisor_shift_id,
        extra_reps=[r.to_rep_dict() for r in Rep.query.filter_by(branch=branch, active=True).all()],
    )

    # هات التقييمات المحفوظة لنفس (اليوم/الفرع/المشرف/شيفته)
    evals = Evaluation.query.filter_by(
        day=day, branch=branch,
        supervisor_id=supervisor_id,
        supervisor_shift_id=supervisor_shift_id,
    ).all()
    evals_by_name = {e.rep_name: e for e in evals}

    session = ShiftSession.query.filter_by(
        day=day, branch=branch,
        supervisor_id=supervisor_id,
        supervisor_shift_id=supervisor_shift_id,
    ).first()

    out = []
    for rep in reps:
        ev = evals_by_name.get(rep["name"])
        out.append({
            "name": rep["name"],
            "type": rep["type"],
            "start": rep["start"],
            "end_hour": D.rep_end_hour(rep) % 24,
            "attendance": ev.attendance if ev else None,
            "reason": ev.reason if ev else None,
            "orders": ev.orders if ev else 0,
            "miss": ev.miss if ev else 0,
            "saved": ev.saved if ev else False,
        })

    return jsonify({
        "reps": out,
        "shift_closed": session is not None,
    })


# ============================================================
# Cortex — جلب بيانات مندوب أوتوماتيك من Amazon Logistics
# ============================================================
@app.route("/api/cortex/request", methods=["POST"])
def api_cortex_request():
    """المشرف داس زرار Cortex -> بنسجل طلب جديد ونرجّع request_uid.
    الفرونت إند بيبعت الـ request_uid + short_name للاسكريبت عبر BroadcastChannel."""
    payload = request.get_json(force=True) or {}
    required = ["branch", "supervisor_id", "supervisor_shift_id", "day", "rep_name"]
    missing = [k for k in required if k not in payload]
    if missing:
        return jsonify({"error": f"missing: {missing}"}), 400

    req = CortexRequest(
        request_uid=str(uuid.uuid4()),
        day=payload["day"],
        branch=payload["branch"],
        supervisor_id=payload["supervisor_id"],
        supervisor_shift_id=payload["supervisor_shift_id"],
        rep_name=payload["rep_name"],
        short_name=C.short_name(payload["rep_name"]),
        status="pending",
    )
    db.session.add(req)
    db.session.commit()

    return jsonify({"ok": True, "request": req.to_dict()})


@app.route("/api/cortex/claim", methods=["POST"])
def api_cortex_claim():
    """الاسكريبت بيعمل polling على ده كل شوية. بيرجع *كل* الطلبات الـ
    pending دفعة واحدة (مش واحد بس)، ويحوّل حالتهم كلهم لـ claimed في
    نفس اللحظة، عشان الاسكريبت يشتغل عليهم واحد ورا التاني من غير ما
    يستنى دورة polling جديدة بين كل طلب والتاني.

    بنستخدم SELECT ... FOR UPDATE SKIP LOCKED عشان لو فاتح أكتر من
    تاب/نافذة Amazon في نفس الوقت، كل تاب ياخد طلبات مختلفة ومحدش
    ياخد نفس الطلب مرتين."""
    reqs = (
        CortexRequest.query.filter_by(status="pending")
        .order_by(CortexRequest.id.asc())
        .with_for_update(skip_locked=True)
        .all()
    )
    if not reqs:
        return jsonify({"ok": True, "requests": []})

    for r in reqs:
        r.status = "claimed"
    db.session.commit()

    return jsonify({"ok": True, "requests": [
        {"request_uid": r.request_uid, "short_name": r.short_name} for r in reqs
    ]})


@app.route("/api/cortex/upload", methods=["POST"])
def api_cortex_upload():
    """الاسكريبت بيرفع ملف الـ CSV هنا بعد ما يحمله من Amazon Logistics.
    multipart/form-data: file=<csv>, request_uid=<uuid>."""
    request_uid = request.form.get("request_uid", "")
    if not request_uid:
        return jsonify({"error": "missing_request_uid"}), 400

    req = CortexRequest.query.filter_by(request_uid=request_uid).first()
    if not req:
        return jsonify({"error": "unknown_request"}), 404

    if "file" not in request.files:
        return jsonify({"error": "missing_file"}), 400

    file_bytes = request.files["file"].read()

    try:
        suggested_orders, raw_row_count, unique_stop_count = C.count_orders_from_csv(file_bytes)
    except ValueError as e:
        req.status = "failed"
        req.error_message = str(e)
        db.session.commit()
        return jsonify({"error": str(e)}), 422

    req.status = "uploaded"
    req.suggested_orders = suggested_orders
    req.raw_row_count = raw_row_count
    req.unique_stop_count = unique_stop_count
    req.uploaded_at = datetime.utcnow()
    db.session.commit()

    return jsonify({"ok": True, "request": req.to_dict()})


@app.route("/api/cortex/status/<request_uid>")
def api_cortex_status(request_uid):
    """الفرونت إند بيعمل polling خفيف على ده لحد ما الطلب يبقى uploaded."""
    req = CortexRequest.query.filter_by(request_uid=request_uid).first()
    if not req:
        return jsonify({"error": "unknown_request"}), 404
    return jsonify({"ok": True, "request": req.to_dict()})


@app.route("/api/cortex/apply", methods=["POST"])
def api_cortex_apply():
    """المشرف قبل الرقم المقترح -> بنحطه في orders بتاع الـ Evaluation
    (بنعيد استخدام نفس شرط قفل الشيفت اللي في /api/evaluate)."""
    payload = request.get_json(force=True) or {}
    request_uid = payload.get("request_uid", "")
    req = CortexRequest.query.filter_by(request_uid=request_uid).first()
    if not req:
        return jsonify({"error": "unknown_request"}), 404
    if req.status != "uploaded" or req.suggested_orders is None:
        return jsonify({"error": "not_ready"}), 409

    session = ShiftSession.query.filter_by(
        day=req.day, branch=req.branch,
        supervisor_id=req.supervisor_id,
        supervisor_shift_id=req.supervisor_shift_id,
    ).first()
    if session:
        return jsonify({"error": "shift_closed"}), 409

    orders_value = payload.get("orders", req.suggested_orders)
    try:
        orders_value = max(0, int(orders_value))
    except (TypeError, ValueError):
        return jsonify({"error": "invalid_orders"}), 400

    ev = Evaluation.query.filter_by(
        day=req.day, branch=req.branch,
        supervisor_id=req.supervisor_id,
        supervisor_shift_id=req.supervisor_shift_id,
        rep_name=req.rep_name,
    ).first()
    if not ev:
        return jsonify({"error": "evaluation_not_found"}), 404

    ev.orders = orders_value
    req.status = "applied"
    req.applied_at = datetime.utcnow()
    db.session.commit()

    return jsonify({"ok": True, "evaluation": ev.to_dict(), "request": req.to_dict()})


# ============================================================
# Save / upsert a single rep evaluation
# ============================================================
@app.route("/api/evaluate", methods=["POST"])
def api_evaluate():
    payload = request.get_json(force=True) or {}

    required = ["branch", "supervisor_id", "supervisor_name", "supervisor_shift_id",
                "day", "rep_name", "rep_type", "rep_start"]
    missing = [k for k in required if k not in payload]
    if missing:
        return jsonify({"error": f"missing: {missing}"}), 400

    # امنع الحفظ لو الشيفت اتقفل بالفعل
    session = ShiftSession.query.filter_by(
        day=payload["day"], branch=payload["branch"],
        supervisor_id=payload["supervisor_id"],
        supervisor_shift_id=payload["supervisor_shift_id"],
    ).first()
    if session:
        return jsonify({"error": "shift_closed"}), 409

    ev = Evaluation.query.filter_by(
        day=payload["day"], branch=payload["branch"],
        supervisor_id=payload["supervisor_id"],
        supervisor_shift_id=payload["supervisor_shift_id"],
        rep_name=payload["rep_name"],
    ).first()

    if not ev:
        ev = Evaluation(
            day=payload["day"], branch=payload["branch"],
            supervisor_id=payload["supervisor_id"],
            supervisor_name=payload["supervisor_name"],
            supervisor_shift_id=payload["supervisor_shift_id"],
            rep_name=payload["rep_name"],
            rep_type=payload["rep_type"],
            rep_start=payload["rep_start"],
        )
        db.session.add(ev)

    if "attendance" in payload:
        ev.attendance = payload["attendance"]
    if "reason" in payload:
        ev.reason = payload["reason"]
    if "orders" in payload:
        ev.orders = max(0, int(payload["orders"]))
    if "miss" in payload:
        ev.miss = max(0, int(payload["miss"]))
    if "saved" in payload:
        ev.saved = bool(payload["saved"])

    db.session.commit()
    return jsonify({"ok": True, "evaluation": ev.to_dict()})


# ============================================================
# Reset a single rep's evaluation for the day
# ============================================================
@app.route("/api/reset", methods=["POST"])
def api_reset():
    payload = request.get_json(force=True) or {}
    ev = Evaluation.query.filter_by(
        day=payload.get("day"), branch=payload.get("branch"),
        supervisor_id=payload.get("supervisor_id"),
        supervisor_shift_id=payload.get("supervisor_shift_id"),
        rep_name=payload.get("rep_name"),
    ).first()
    if ev:
        db.session.delete(ev)
        db.session.commit()
    return jsonify({"ok": True})


# ============================================================
# Close shift -> lock evaluations + return CSV export
# ============================================================
@app.route("/api/close_shift", methods=["POST"])
def api_close_shift():
    payload = request.get_json(force=True) or {}
    required = ["branch", "supervisor_id", "supervisor_name", "supervisor_shift_id", "day"]
    missing = [k for k in required if k not in payload]
    if missing:
        return jsonify({"error": f"missing: {missing}"}), 400

    existing = ShiftSession.query.filter_by(
        day=payload["day"], branch=payload["branch"],
        supervisor_id=payload["supervisor_id"],
        supervisor_shift_id=payload["supervisor_shift_id"],
    ).first()
    if not existing:
        session = ShiftSession(
            day=payload["day"], branch=payload["branch"],
            supervisor_id=payload["supervisor_id"],
            supervisor_name=payload["supervisor_name"],
            supervisor_shift_id=payload["supervisor_shift_id"],
        )
        db.session.add(session)

    evals = Evaluation.query.filter_by(
        day=payload["day"], branch=payload["branch"],
        supervisor_id=payload["supervisor_id"],
        supervisor_shift_id=payload["supervisor_shift_id"],
    ).all()
    for ev in evals:
        ev.closed = True

    db.session.commit()
    return jsonify({"ok": True})


# ============================================================
# Fines (الغرامات)
# ============================================================
@app.route("/api/fines")
def api_fines_list():
    branch = request.args.get("branch", "")
    supervisor_id = request.args.get("supervisor_id", "")
    supervisor_shift_id = request.args.get("shift_id", "")
    day = request.args.get("day", "")

    if not all([branch, supervisor_id, supervisor_shift_id, day]):
        return jsonify({"error": "missing params"}), 400

    fines = Fine.query.filter_by(
        day=day, branch=branch,
        supervisor_id=supervisor_id,
        supervisor_shift_id=supervisor_shift_id,
    ).order_by(Fine.created_at.desc()).all()

    return jsonify({"fines": [f.to_dict() for f in fines]})


@app.route("/api/fines", methods=["POST"])
def api_fines_add():
    payload = request.get_json(force=True) or {}
    required = ["branch", "supervisor_id", "supervisor_name", "supervisor_shift_id",
                "day", "rep_name", "amount"]
    missing = [k for k in required if k not in payload]
    if missing:
        return jsonify({"error": f"missing: {missing}"}), 400

    session = ShiftSession.query.filter_by(
        day=payload["day"], branch=payload["branch"],
        supervisor_id=payload["supervisor_id"],
        supervisor_shift_id=payload["supervisor_shift_id"],
    ).first()
    if session:
        return jsonify({"error": "shift_closed"}), 409

    try:
        amount = float(payload["amount"])
    except (TypeError, ValueError):
        return jsonify({"error": "invalid_amount"}), 400
    if amount <= 0:
        return jsonify({"error": "invalid_amount"}), 400

    fine = Fine(
        day=payload["day"], branch=payload["branch"],
        supervisor_id=payload["supervisor_id"],
        supervisor_name=payload["supervisor_name"],
        supervisor_shift_id=payload["supervisor_shift_id"],
        rep_name=payload["rep_name"],
        amount=amount,
        reason=(payload.get("reason") or "").strip(),
    )
    db.session.add(fine)
    db.session.commit()
    return jsonify({"ok": True, "fine": fine.to_dict()})


@app.route("/api/fines/<int:fine_id>", methods=["DELETE"])
def api_fines_delete(fine_id):
    fine = Fine.query.get(fine_id)
    if fine:
        db.session.delete(fine)
        db.session.commit()
    return jsonify({"ok": True})


# ============================================================
# Export -> ملف Excel منظم بشيتين: التقييمات + الغرامات
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
SUBTITLE_FONT = Font(name="Calibri", size=11, color="475569")
THIN = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="right", vertical="center", wrap_text=True)

ROW_FILL_EVEN = PatternFill("solid", fgColor="F1F5F9")
ROW_FILL_PRESENT = PatternFill("solid", fgColor="DCFCE7")
ROW_FILL_LATE = PatternFill("solid", fgColor="FEF3C7")
ROW_FILL_ABSENT = PatternFill("solid", fgColor="FEE2E2")
FINE_FILL = PatternFill("solid", fgColor="FEE2E2")


def _sheet_header(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = SUBTITLE_FONT
    c2.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 20
    ws.sheet_view.rightToLeft = True


def _write_table_header(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = CELL_BORDER
    ws.row_dimensions[row].height = 22


# ============================================================
# Export -> ملف Excel بنفس منظر الشيت الأصلي (full/part + الاسم + start)
# + الحضور + الأوردرات/MISS/الغرامات جنب كل مندوب
# ============================================================
NAVY = "1E293B"
YELLOW = "FBBF24"
PURPLE_DARK = "6C3FA8"
PURPLE_MED = "9B59B6"
PURPLE_ROW_A = "F3ECFA"
PURPLE_ROW_B = "EDE1F8"
PART_ROW_A = "F6EBFB"
PART_ROW_B = "F0E0F9"

GREEN = "16A34A"
GREEN_LIGHT = "DCFCE7"
AMBER = "D97706"
AMBER_LIGHT = "FEF3C7"
RED = "DC2626"
RED_LIGHT = "FEE2E2"
GRAY_LIGHT = "F1F5F9"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=17, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=11, color="64748B")
THIN = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)

FINE_FILL = PatternFill("solid", fgColor=RED_LIGHT)


def _sheet_header(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = SUBTITLE_FONT
    c2.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 20
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False


def _write_table_header(ws, row, headers, fill=None, font=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font or HEADER_FONT
        c.fill = fill or HEADER_FILL
        c.alignment = CENTER
        c.border = CELL_BORDER
    ws.row_dimensions[row].height = 24


@app.route("/api/export_csv")
def api_export_csv():
    branch = request.args.get("branch", "")
    supervisor_id = request.args.get("supervisor_id", "")
    supervisor_shift_id = request.args.get("shift_id", "")
    day = request.args.get("day", "")

    reps = D.get_overlapping_reps(
        branch, supervisor_shift_id,
        extra_reps=[r.to_rep_dict() for r in Rep.query.filter_by(branch=branch, active=True).all()],
    )

    evals = Evaluation.query.filter_by(
        day=day, branch=branch,
        supervisor_id=supervisor_id,
        supervisor_shift_id=supervisor_shift_id,
    ).all()
    eval_by_name = {e.rep_name: e for e in evals}

    fines = Fine.query.filter_by(
        day=day, branch=branch,
        supervisor_id=supervisor_id,
        supervisor_shift_id=supervisor_shift_id,
    ).order_by(Fine.created_at).all()
    fines_by_name = {}
    for f in fines:
        fines_by_name.setdefault(f.rep_name, []).append(f)

    branch_label = D.BRANCHES.get(branch, {}).get("label", branch)
    shift = D.get_supervisor_shift(supervisor_shift_id)
    shift_label = shift["label"] if shift else supervisor_shift_id
    sup_name = evals[0].supervisor_name if evals else (fines[0].supervisor_name if fines else "")
    subtitle = f"الفرع: {branch_label}      المشرف: {sup_name}      الشيفت: {shift_label}      اليوم: {day}"

    reps_sorted = sorted(reps, key=lambda r: (0 if r["type"] == "full" else 1, r["start"]))

    wb = Workbook()

    # ================= Sheet 1: المندوبين =================
    ws = wb.active
    ws.title = "المندوبين"
    headers = ["نوع الدوام", "اسم المندوب", "بداية الشيفت", "الحضور", "الأوردرات", "MISS", "الغرامات (جنيه)"]
    ncols = len(headers)
    _sheet_header(ws, "🚚  BANKAI — جدول المندوبين", subtitle, ncols)

    header_row = 4
    _write_table_header(ws, header_row, headers)

    dv = DataValidation(type="list", formula1='"full time,part time"', allow_blank=True)
    ws.add_data_validation(dv)

    att_map = {"present": "✓ حضر", "late": "⏱ تأخير", "absent": "✕ غياب"}
    att_colors = {
        "present": (GREEN, GREEN_LIGHT),
        "late": (AMBER, AMBER_LIGHT),
        "absent": (RED, RED_LIGHT),
    }

    r = header_row + 1
    full_count = part_count = 0
    total_present = total_late = total_absent = 0

    for i, rep in enumerate(reps_sorted):
        ev = eval_by_name.get(rep["name"])
        rep_fines = fines_by_name.get(rep["name"], [])
        fine_total = sum(f.amount for f in rep_fines)

        orders = ev.orders if ev else 0
        miss = ev.miss if ev else 0
        attendance = ev.attendance if ev else None
        is_full = rep["type"] == "full"

        if attendance == "present": total_present += 1
        elif attendance == "late": total_late += 1
        elif attendance == "absent": total_absent += 1

        ws.row_dimensions[r].height = 20

        # عمود 1: بادچ نوع الدوام (dropdown)
        c1 = ws.cell(row=r, column=1, value="full time" if is_full else "part time")
        c1.font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
        c1.fill = PatternFill("solid", fgColor=PURPLE_DARK if is_full else PURPLE_MED)
        c1.alignment = CENTER
        c1.border = CELL_BORDER
        dv.add(c1)

        # عمود 2: الاسم
        c2 = ws.cell(row=r, column=2, value=rep["name"])
        c2.font = Font(name="Calibri", size=11, bold=True, color="1E293B")
        c2.alignment = RIGHT_ALIGN
        c2.border = CELL_BORDER

        # عمود 3: بداية الشيفت
        hr = rep['start'] % 24
        period = "ص" if hr < 12 else "م"
        hr12 = hr if 1 <= hr <= 12 else (12 if hr == 0 else hr - 12)
        c3 = ws.cell(row=r, column=3, value=f"{hr12}:00 {period}")
        c3.font = Font(name="Calibri", size=10.5, color="475569")
        c3.alignment = CENTER
        c3.border = CELL_BORDER

        # عمود 4: الحضور
        if attendance:
            label = att_map.get(attendance, "-")
            fg, bg = att_colors.get(attendance, ("475569", GRAY_LIGHT))
        else:
            label, fg, bg = "— لم يُسجل —", "94A3B8", GRAY_LIGHT
        c4 = ws.cell(row=r, column=4, value=label)
        c4.font = Font(name="Calibri", size=10.5, bold=True, color=fg)
        c4.alignment = CENTER
        c4.border = CELL_BORDER
        c4.fill = PatternFill("solid", fgColor=bg)

        # عمود 5: الأوردرات
        c5 = ws.cell(row=r, column=5, value=orders)
        c5.alignment = CENTER
        c5.border = CELL_BORDER
        c5.font = Font(name="Calibri", size=11, bold=orders > 0)

        # عمود 6: MISS
        c6 = ws.cell(row=r, column=6, value=miss)
        c6.alignment = CENTER
        c6.border = CELL_BORDER
        c6.font = Font(name="Calibri", size=11, bold=miss > 0, color=RED if miss > 0 else "1E293B")

        # عمود 7: الغرامات
        c7 = ws.cell(row=r, column=7, value=fine_total)
        c7.alignment = CENTER
        c7.border = CELL_BORDER
        c7.number_format = '#,##0.00 "ج.م"'
        c7.font = Font(name="Calibri", size=11, bold=fine_total > 0, color=RED if fine_total > 0 else "1E293B")
        if fine_total > 0:
            c7.fill = FINE_FILL

        # خلفية الصف (تبادل حسب النوع)
        if is_full:
            row_fill = PatternFill("solid", fgColor=PURPLE_ROW_A if i % 2 == 0 else PURPLE_ROW_B)
        else:
            row_fill = PatternFill("solid", fgColor=PART_ROW_A if i % 2 == 0 else PART_ROW_B)
        for c in (c2, c3, c5, c6):
            c.fill = row_fill

        if is_full:
            full_count += 1
        else:
            part_count += 1
        r += 1

    if reps_sorted:
        total_row = r
        ws.row_dimensions[total_row].height = 24
        label = f"الإجمالي  —  {full_count} دوام كامل  /  {part_count} دوام جزئي   |   حضر {total_present}  •  تأخير {total_late}  •  غياب {total_absent}"
        ws.cell(row=total_row, column=1, value=label).font = Font(bold=True, size=10, color="FFFFFF")
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=NAVY)

        c_orders = ws.cell(row=total_row, column=5, value=f"=SUM(E{header_row+1}:E{r-1})")
        c_miss = ws.cell(row=total_row, column=6, value=f"=SUM(F{header_row+1}:F{r-1})")
        c_fines = ws.cell(row=total_row, column=7, value=f"=SUM(G{header_row+1}:G{r-1})")
        c_fines.number_format = '#,##0.00 "ج.م"'
        for c in (c_orders, c_miss, c_fines):
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = CENTER
        for col in range(1, ncols + 1):
            ws.cell(row=total_row, column=col).border = CELL_BORDER
    else:
        ws.cell(row=header_row + 1, column=1, value="لا يوجد مندوبين لهذا الشيفت")
        ws.merge_cells(start_row=header_row+1, start_column=1, end_row=header_row+1, end_column=ncols)
        ws.cell(row=header_row+1, column=1).alignment = CENTER

    col_widths = [13, 30, 14, 14, 12, 8, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row+1}"

    # ================= Sheet 2: الغرامات (تفصيلي) =================
    ws2 = wb.create_sheet("الغرامات")
    headers2 = ["#", "اسم المندوب", "المبلغ (جنيه)", "السبب"]
    ncols2 = len(headers2)
    _sheet_header(ws2, "💰  BANKAI — تقرير الغرامات", subtitle, ncols2)

    header_row2 = 4
    _write_table_header(ws2, header_row2, headers2)

    r2 = header_row2 + 1
    for idx, f in enumerate(fines, start=1):
        ws2.row_dimensions[r2].height = 20
        values = [idx, f.rep_name, f.amount, f.reason or "-"]
        row_fill = PatternFill("solid", fgColor=RED_LIGHT if idx % 2 == 1 else "FEF2F2")
        for ci, v in enumerate(values, start=1):
            c = ws2.cell(row=r2, column=ci, value=v)
            c.border = CELL_BORDER
            c.alignment = CENTER if ci != 4 else RIGHT_ALIGN
            c.font = Font(name="Calibri", size=11, bold=(ci == 2))
            if ci == 3:
                c.number_format = '#,##0.00 "ج.م"'
            c.fill = row_fill
        r2 += 1

    if fines:
        total_row2 = r2
        ws2.cell(row=total_row2, column=1, value="إجمالي الغرامات").font = Font(bold=True, color="FFFFFF")
        ws2.merge_cells(start_row=total_row2, start_column=1, end_row=total_row2, end_column=2)
        ws2.cell(row=total_row2, column=1).alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c_total = ws2.cell(row=total_row2, column=3, value=f"=SUM(C{header_row2+1}:C{r2-1})")
        c_total.number_format = '#,##0.00 "ج.م"'
        for c in (ws2.cell(row=total_row2, column=1), c_total, ws2.cell(row=total_row2, column=4)):
            c.font = Font(bold=True, color="FFFFFF")
            c.border = CELL_BORDER
            c.fill = PatternFill("solid", fgColor=RED)
            c.alignment = CENTER
    else:
        ws2.cell(row=header_row2 + 1, column=1, value="لا توجد غرامات مسجلة لهذا الشيفت")
        ws2.merge_cells(start_row=header_row2+1, start_column=1, end_row=header_row2+1, end_column=ncols2)
        ws2.cell(row=header_row2+1, column=1).alignment = CENTER

    col_widths2 = [6, 28, 16, 42]
    for i, w in enumerate(col_widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = f"A{header_row2+1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"bankai-{branch}-{supervisor_id}-{day}-{supervisor_shift_id}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
